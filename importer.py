from openpyxl import load_workbook
from db import add_question


def import_xlsx(path):
    wb = load_workbook(path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise Exception("Excel faylda maʼlumot yo‘q")

    header = rows[0]
    required = [
        "subject", "block", "difficulty",
        "question", "A", "B", "C", "D", "correct"
    ]

    for r in required:
        if r not in header:
            raise Exception(f"Ustun yetishmayapti: {r}")

    idx = {name: header.index(name) for name in required}

    count = 0

    for i, row in enumerate(rows[1:], start=2):
        try:
            subject = row[idx["subject"]]
            block = row[idx["block"]]
            difficulty = row[idx["difficulty"]]
            question = row[idx["question"]]
            A = row[idx["A"]]
            B = row[idx["B"]]
            C = row[idx["C"]]
            D = row[idx["D"]]
            correct = row[idx["correct"]]

            if correct not in ("A", "B", "C", "D"):
                raise Exception("correct faqat A/B/C/D bo‘lishi kerak")

            add_question(
                subject,
                block,
                difficulty,
                question,
                A, B, C, D,
                correct
            )
            count += 1

        except Exception as e:
            raise Exception(f"{i}-qatorda xato: {e}")

    return count
